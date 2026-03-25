# Excel Indent Function

import pandas as pd
import os
import re
import sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def indent_function(excel_file, heading_column, indent_column):
    """
    Reads an Excel file, applies native Excel indentation to a specified column
    based on values in another column, and saves the modified workbook
    to a new Excel file with an '_indented' suffix in the same directory.

    Args:
        excel_file (str): The full path to the Excel file to manipulate.
        heading_column (str or int): The name or index of the column
                                     containing the numerical indent information.
        indent_column (str or int): The name or index of the column
                                     whose cells are to be indented.

    Returns:
        tuple: A tuple containing:
            - str: The path to the newly created indented Excel file, or an empty string if an error occurred.
            - list: A list of strings containing all output messages from the function.
    """
    output = []
    excel_file2 = "" 

    try:
        # Load workbook using openpyxl to manipulate formatting directly
        wb = load_workbook(excel_file)
        ws = wb.active # Operates on the primary/active sheet
        output.append("Successfully read file with openpyxl.")
    except FileNotFoundError:
        output.append(f"Error: File '{excel_file}' not found.")
        return "", output
    except Exception as e:
        output.append(f"Error reading Excel file '{excel_file}': {e}")
        return "", output

    # Extract headers from the first row to map column names to openpyxl's 1-based indices
    headers = [cell.value for cell in ws[1]]

    def get_col_idx(col_identifier):
        """Helper to convert pandas 0-based index or string to openpyxl 1-based index."""
        if isinstance(col_identifier, str):
            if col_identifier in headers:
                return headers.index(col_identifier) + 1
        elif isinstance(col_identifier, int):
            # Convert pandas 0-based column index to openpyxl 1-based index
            if 0 <= col_identifier < len(headers):
                return col_identifier + 1
        return -1

    # Identify the column to be used for number of indents
    heading_col_idx = get_col_idx(heading_column)
    if heading_col_idx == -1:
        output.append(f"Error: Heading column '{heading_column}' not found or out of bounds.")
        return "", output

    # Identify the column to be indented
    indent_col_idx = get_col_idx(indent_column)
    if indent_col_idx == -1:
        output.append(f"Error: Indent column '{indent_column}' not found or out of bounds.")
        return "", output

    # Iterate through rows starting from row 2 (skipping header)
    for row in range(2, ws.max_row + 1):
        heading_cell = ws.cell(row=row, column=heading_col_idx)
        indent_cell = ws.cell(row=row, column=indent_col_idx)

        try:
            if heading_cell.value is not None:
                # Get the numerical indent level
                indent_level = int(float(heading_cell.value))
                
                # Apply native Excel indentation using openpyxl Alignment
                # We copy existing alignment properties so we don't overwrite things like wrap_text
                current_alignment = indent_cell.alignment
                if current_alignment:
                    indent_cell.alignment = Alignment(
                        horizontal=current_alignment.horizontal,
                        vertical=current_alignment.vertical,
                        text_rotation=current_alignment.text_rotation,
                        wrap_text=current_alignment.wrap_text,
                        shrink_to_fit=current_alignment.shrink_to_fit,
                        indent=indent_level
                    )
                else:
                    indent_cell.alignment = Alignment(indent=indent_level)

        except (ValueError, TypeError):
            output.append(f"Warning: Row {row}: Invalid or missing numeric value in heading column. Skipping indentation for this row.")

    # Save the changes
    suffix = "_indented"
    base_name, ext = os.path.splitext(excel_file)
    excel_file2 = os.path.join(os.path.dirname(excel_file), f"{os.path.basename(base_name)}{suffix}{ext}")

    try:
        wb.save(excel_file2)
        output.append(f"File '{excel_file2}' updated successfully.")
        return excel_file2, output 
    except Exception as e:
        output.append(f"Error saving Excel file '{excel_file2}': {e}")
        return "", output


def calculate_indents_and_save_new_excel(excel_file_name: str, heading_column: str = 'Heading') -> tuple:
    """
    Reads an Excel file, calculates the number of indents for each entry in a specified heading column,
    appends these indents as a new column to the DataFrame, and then saves
    the modified DataFrame to a new Excel file with a '_new' suffix in the same directory as the input file.

    The indent calculation logic is as follows:
    - For numbered headings (e.g., '1. Title', '1.1 Subtitle', '1.1.1 Sub-Subtitle'):
      The indent is determined by counting the number of dots in the numbering prefix.
      Example:
        '1. Title' -> 0 indents
        '1.1 Subtitle' -> 1 indent
        '1.1.1 Sub-Subtitle' -> 2 indents
    - For non-numbered text (e.g., 'Requirement Text'):
      The indent is one more than the indent of the last encountered numbered heading.

    Args:
        excel_file_name (str): The full path to the Excel file (e.g., 'C:/Users/User/Documents/my_data.xlsx').
        heading_column (str, optional): The name of the column in the Excel file
                                        that contains the headings. Defaults to 'Heading'.

    Returns:
        tuple: A tuple containing:
            - str: The full path to the newly created Excel file, or an empty string if an error occurred.
            - int: The column index of the newly added 'Calculated Indents' column, or -1 if not created.
            - int: The column index of the 'Heading' column, or -1 if not found.
            - list: A list of strings containing all output messages from the function.
    """
    output = []
    df = None
    output_excel_file_name = ""
    new_column_index = -1
    heading_column_index = -1

    try:
        df = pd.read_excel(excel_file_name)
        output.append(f"Successfully read file.")
    except FileNotFoundError:
        output.append(f"Error: File '{excel_file_name}' not found.")
        return "", new_column_index, heading_column_index, output
    except Exception as e:
        output.append(f"Error reading Excel file '{excel_file_name}': {e}")
        return "", new_column_index, heading_column_index, output 

    if heading_column not in df.columns:
        output.append(f"Error: '{heading_column}' column not found in the Excel file.")
        return "", new_column_index, heading_column_index, output 

    calculated_indents = []
    last_numbered_heading_indent = -1

    full_numeric_prefix_pattern = re.compile(r'^(\d+(\.\d+)*)')

    for index, row in df.iterrows():
        heading = str(row[heading_column]).strip()
        full_numeric_prefix_match = full_numeric_prefix_pattern.match(heading)

        if full_numeric_prefix_match:
            prefix = full_numeric_prefix_match.group(1)
            current_indent = prefix.count('.')
            last_numbered_heading_indent = current_indent
            calculated_indents.append(current_indent)
        else:
            if last_numbered_heading_indent == -1:
                calculated_indents.append(0)
            else:
                calculated_indents.append(last_numbered_heading_indent + 1)

    df['Calculated Indents'] = calculated_indents

    try:
        new_column_index = df.columns.get_loc('Calculated Indents')
        heading_column_index = df.columns.get_loc(heading_column)
    except KeyError as e:
        output.append(f"Error getting column index: {e}. This should not happen after successful column creation/check.")

    # Construct the output file name to be in the same directory as the input
    base_name, ext = os.path.splitext(excel_file_name)
    output_excel_file_name = os.path.join(os.path.dirname(excel_file_name), f"{os.path.basename(base_name)}_new{ext}")

    try:
        df.to_excel(output_excel_file_name, index=False)
        output.append(f"Successfully saved results to: '{output_excel_file_name}'.")
        return output_excel_file_name, new_column_index, heading_column_index, output
    except Exception as e:
        output.append(f"Error saving Excel file '{output_excel_file_name}': {e}")
        return "", new_column_index, heading_column_index, output